import openpyxl
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font

from .models import (
    Course,
    CourseEnrollment,
    CourseInstructor,
    CourseLesson,
    CourseRating,
    CourseTechnology,
)


@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    search_fields = ("title", "slug", "instructor__name")


@admin.register(CourseTechnology)
class TechnologyAdmin(admin.ModelAdmin):
    search_fields = ("name",)


@admin.register(CourseInstructor)
class CourseInstructorAdmin(admin.ModelAdmin):
    search_fields = ("name", "course__title")


@admin.register(CourseLesson)
class CourseLessonAdmin(admin.ModelAdmin):
    search_fields = ("title", "course__title")


@admin.register(CourseRating)
class CourseRatingAdmin(admin.ModelAdmin):
    search_fields = ("user__username", "course__title")


@admin.register(CourseEnrollment)
class CourseEnrollmentAdmin(admin.ModelAdmin):
    search_fields = ("course__title", "user__email")
    actions = ["download_enrollments_excel"]

    def download_enrollments_excel(self, request, queryset):

        workbook = openpyxl.Workbook()
        worksheet = workbook.active or workbook.create_sheet()
        worksheet.title = "Course Enrollments"

        headers = ["Student ID", "Name", "Phone", "Email", "Course", "Year"]

        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        if queryset.exists():
            enrollments = queryset.select_related("user", "course")
        else:
            enrollments = CourseEnrollment.objects.select_related(
                "user", "course"
            ).all()

        row = 2
        for enrollment in enrollments:
            worksheet.cell(row=row, column=1, value=enrollment.user.pk)
            worksheet.cell(
                row=row,
                column=2,
                value=enrollment.user.name or enrollment.user.username,
            )
            worksheet.cell(row=row, column=3, value=enrollment.user.mobile or "")
            worksheet.cell(row=row, column=4, value=enrollment.user.email)
            worksheet.cell(row=row, column=5, value=enrollment.course.title)
            worksheet.cell(
                row=row,
                column=6,
                value=enrollment.user.year,
            )
            row += 1

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="course_enrollments_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )

        workbook.save(response)
        return response
